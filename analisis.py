
import os.path, time
from PIL import Image
from PIL.ExifTags import TAGS

dir = os.getcwd()
file = os.listdir()


dir = os.getcwd() + "\\" + file[0]
date_create_photo = {}

type(date_create_photo)

date_create_photo["IMG_20230714_173703.jpg"]


date_create_photo

value_KWh = [496398,
        496398,
        496398,
        496447,
        496505,
        496512,
        496525,
        496548,
        496576,
        496598,
        496625,
        496647,
        496670,
        496711,
        496731,
        496789,
        496828,
        496828,
        496835,
        496837,
        496837,
        496852,
        496854,
        476876,
        496919,
        496919,
        496946,
        496977,
        496977,
        497009,
        497026,
        497051,
        497071,
        497098,
        497174,
        497197,
        497200,
        497209,
        497237,
        497262,
        497289,
        497312,
        497345,
        497377,
        497393,
        497422,
        497427,
        497458,
        497461,
        497539,
        497543,
        497589,
        497593,
        497655,
        497657,
        497687,
        497705,
        497732,
        497766,
        497786,
        497801,
        497827,
        497845,
        497894,
        497911,
        497972,
        497973,
        498005,
        498009,
        498034,
        498052,
        498080,
        498105,
        498133,
        498149,
        498191,
        498195,
        498245,
        498249,
        498261,
        498345,
        498399,
        498424,
        498447,
        496709,
        496709,
]

type(file)

""" Extracting data create file
for i in file:
    print("fecha creación: {}".format(time.ctime(os.path.getctime(i))))

"""

for i in file:
    # open the image
    image = Image.open(i)
    
    # extracting the exif metadata
    exifdata = image.getexif()
    
    j=0

    # looping through all the tags present in exifdata
    for tagid in exifdata:
        
        # getting the tag name instead of tag id
        tagname = TAGS.get(tagid, tagid)
    
        # passing the tagid to get its respective value
        value = exifdata.get(tagid)
        
        # printing the final result
        print(f"{tagname:25}: {value}")

        if tagname == "DateTime":
            date_create_photo[i] = value
            date_create_photo[i] = value_KWh[j]
    j=j+1


for i in range(0, len(value_KWh)):
    date_create_photo["KWh"] = value_KWh[i]













from twilio.twiml.messaging_response import MessagingResponse
import pandas as pd
from openpyxl import load_workbook

def process_and_store_message(sender, message):
    # Procesar la solicitud (aquí puedes agregar tu lógica de procesamiento)
    response_text = f"Gracias por tu solicitud, {sender.split(':')[1]}."

    # Almacenar la solicitud en Excel
    data = {'Remitente': [sender], 'Mensaje': [message]}
    df = pd.DataFrame(data)

    # Agregar a un archivo Excel existente o crear uno nuevo
    excel_file = 'solicitudes.xlsx'
    try:
        book = load_workbook(excel_file)
        writer = pd.ExcelWriter(excel_file, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        writer.save()
    except FileNotFoundError:
        df.to_excel(excel_file, index=False)

    return response_text

# Simulación de recibir mensaje de WhatsApp
sender = "whatsapp:+1234567890"
message = "¡Hola! Quiero hacer una solicitud."

response = process_and_store_message(sender, message)
print(response)





































496398,
496398,
496398,
496447,
496505,
496512,
496525,
496548,
496576,
496598,
496625,
496647,
496670,
496711,
496731,
496789,
496828,
496828,
496835,
496837,
496837,
496852,
496854,
476876,
496919,
496919,
496946,
496977,
496977,
497009,
497026,
497051,
497071,
497098,
497174,
497197,
497200,
497209,
497237,
497262,
497289,
497312,
497345,
497377,
497393,
497422,
497427,
497458,
497461,
497539,
497543,
497589,
497593,
497655,
497657,
497687,
497705,
497732,
497766,
497786,
497801,
497827,
497845,
497894,
497911,
497972,
497973,
498005,
498009,
498034,
498052,
498080,
498105,
498133,
498149,
498191,
498195,
498245,
498249,
498261,
498345,
498399,
498424,
498447,
496709,
496709,
